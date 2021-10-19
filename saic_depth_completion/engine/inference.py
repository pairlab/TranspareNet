import os
import time
import datetime
import torch
import torch.nn.functional as nnf
from tqdm import tqdm

import matplotlib.pyplot as plt

from saic_depth_completion.utils.meter import AggregatedMeter
from saic_depth_completion.utils.meter import Statistics as LossMeter
from saic_depth_completion.utils import visualize

import numpy as np
import OpenEXR
import Imath
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

def exr_saver(EXR_PATH, ndarr, ndim=3):
    '''Saves a numpy array as an EXR file with HALF precision (float16)
    Args:
        EXR_PATH (str): The path to which file will be saved
        ndarr (ndarray): A numpy array containing img data
        ndim (int): The num of dimensions in the saved exr image, either 3 or 1.
                        If ndim = 3, ndarr should be of shape (height, width) or (3 x height x width),
                        If ndim = 1, ndarr should be of shape (height, width)
    Returns:
        None
    '''
    if ndim == 3:
        # Check params
        if len(ndarr.shape) == 2:
            # If a depth image of shape (height x width) is passed, convert into shape (3 x height x width)
            ndarr = np.stack((ndarr, ndarr, ndarr), axis=0)

        if ndarr.shape[0] != 3 or len(ndarr.shape) != 3:
            raise ValueError(
                'The shape of the tensor should be (3 x height x width) for ndim = 3. Given shape is {}'.format(
                    ndarr.shape))

        # Convert each channel to strings
        Rs = ndarr[0, :, :].astype(np.float16).tostring()
        Gs = ndarr[1, :, :].astype(np.float16).tostring()
        Bs = ndarr[2, :, :].astype(np.float16).tostring()

        # Write the three color channels to the output file
        HEADER = OpenEXR.Header(ndarr.shape[2], ndarr.shape[1])
        half_chan = Imath.Channel(Imath.PixelType(Imath.PixelType.HALF))
        HEADER['channels'] = dict([(c, half_chan) for c in "RGB"])

        out = OpenEXR.OutputFile(EXR_PATH, HEADER)
        out.writePixels({'R': Rs, 'G': Gs, 'B': Bs})
        out.close()
    elif ndim == 1:
        # Check params
        if len(ndarr.shape) != 2:
            raise ValueError(('The shape of the tensor should be (height x width) for ndim = 1. ' +
                              'Given shape is {}'.format(ndarr.shape)))

        # Convert each channel to strings
        Rs = ndarr[:, :].astype(np.float16).tostring()

        # Write the color channel to the output file
        HEADER = OpenEXR.Header(ndarr.shape[1], ndarr.shape[0])
        half_chan = Imath.Channel(Imath.PixelType(Imath.PixelType.HALF))
        HEADER['channels'] = dict([(c, half_chan) for c in "R"])

        out = OpenEXR.OutputFile(EXR_PATH, HEADER)
        out.writePixels({'R': Rs})
        out.close()

def inference(
        model, test_loaders, metrics, pccPred=None, save_dir="", logger=None, file_names = None, visualize_pics = False,store_computed_metrics = False, prefix='transparenet10k'
):

    model.eval()
    metrics_meter = AggregatedMeter(metrics, maxlen=20)
    for subset, loader in test_loaders.items():
        idx = 0
        logger.info(
            "Inference: subset -- {}. Total number of batches: {}.".format(subset, len(loader))
        )
        if file_names is not None:
            split_files = file_names[subset].color_name
            arrangement_path, _ = os.path.split(split_files[0])
            data_path, _ = os.path.split(os.path.split(arrangement_path)[0])
        metrics_meter.reset()
        # loop over dataset
        rmse_temp = list()
        for batch in tqdm(loader):
            
            if pccPred is not None:
                batch = pccPred(batch)
            batch = model.preprocess(batch)
            pred = model(batch)

            with torch.no_grad():
                arrangement_path, _ = os.path.split(split_files[idx])
                post_pred = model.postprocess(pred)
                if save_dir:
                    if file_names is not None:

                        if visualize_pics:
                            B = batch["color"].shape[0]
                            assert B == 1, f'Ensure batchsize is 1, current batchsize is {B}'
                            for it in range(B):
                                file_name = split_files[idx]
                                fig = visualize.figure(
                                    batch["color"][it], batch["raw_depth"][it],
                                    batch["mask"][it], batch["gt_depth"][it],
                                    post_pred[it], close=True
                                )
                                data_folder_path, rgb_name = os.path.split(file_name)
                                exr_saver(EXR_PATH = os.path.join(data_folder_path,f'{prefix}_depth.exr'), ndarr = post_pred[it].detach().cpu()[0].numpy(), ndim=1)
                                fig.savefig(
                                    os.path.join(data_folder_path,f'{prefix}_result_compare.png'), dpi=fig.dpi
                                )
                                path_save = os.path.join(data_folder_path,f'{prefix}_result_compare.png')
                                print(f'saving fig to {path_save}')
                                #idx += 1
                    else:
                        if visualize_pics:
                            B = batch["color"].shape[0]
                            assert B == 1, f'Ensure batchsize is 1, current batchsize is {B}'
                            for it in range(B):
                                fig = visualize.figure(
                                    batch["color"][it], batch["raw_depth"][it],
                                    batch["mask"][it], batch["gt_depth"][it],
                                    post_pred[it], close=True
                                )
                                fig.savefig(
                                    os.path.join(save_dir, "result_{}.png".format(idx)), dpi=fig.dpi
                                )

                                #idx += 1
                # Reshape tensors to proper dimension
                # print(batch["mask"])
                # print(post_pred.shape)
                post_pred = nnf.interpolate(post_pred, size=(144,256))
                batch["gt_depth"] = nnf.interpolate(batch["gt_depth"], size=(144,256))

                # NaN goes to zero in gt_depth
                batch['gt_depth'][batch['gt_depth']!= batch['gt_depth']] = 0
                # batch["gt_depth"] = torch.nan_to_num(batch['gt_depth'],nan = 0.0)
                # print(batch['gt_depth'])
                mask_valid_region = (batch['gt_depth'] > 0.01).byte()
                try:
                    mask_seg = (nnf.interpolate(batch["gt_mask"], size=(144,256)) <=0.01).byte()
                except:
                    raise ValueError('no gt_mask attribute')
                # print('mask valid region', torch.sum(mask_valid_region))
                # print('mask seg',torch.sum(mask_seg))
                mask_bool = mask_valid_region.__and__(mask_seg)
                # print('masksksk', mask_bool, mask_bool.size())

                # print('max_gt', torch.max(batch["gt_depth"][mask_bool]))
                # print('mean_gt', torch.mean(batch["gt_depth"][mask_bool]))
                # print('max_pred', torch.max(post_pred[mask_bool]))
                # print('mean_pred', torch.mean(post_pred[mask_bool]))
                # print('mask bool',torch.sum(mask_bool))
                # print('mask_valid_region', torch.sum(mask_valid_region))
                # mask_bool = torch.logical_and(mask_valid_region,mask_seg).byte()
                # print('number of elements',batch["gt_depth"][mask_bool].numel())
                rmse_temp.append(((batch["gt_depth"][mask_bool] - post_pred[mask_bool])**2).mean().sqrt().cpu().detach().numpy())
                # try:
                # print('pred dims', post_pred,post_pred.size())
                # metrics_meter.update(post_pred[mask_bool], batch["gt_depth"][mask_bool])
                # except:
                #     print('ERRRORR IN METER')
                #     continue
                try:
                    metrics_meter.update(post_pred[mask_bool], batch["gt_depth"][mask_bool])
                except:
                    print('ERRRORR IN METER')
                    continue
                # Update excel spreadsheet with metrics
                if store_computed_metrics:
                    # Split path:
                    headers = ['path']+ list(metrics.keys())
                    # Check if excel spreadsheet already exists
                    if os.path.exists(os.path.join(data_path,subset+'.xlsx')):
                        # Append to spreadsheet
                        workbook_name = os.path.join(data_path,subset+'.xlsx')
                        wb = load_workbook(workbook_name)
                        page = wb.active

                        # New data to write:
                        data_to_write = [[arrangement_path] + [float(metrics_meter.meters[header_name].stats.enum[-1]) for header_name in headers[1:]]]
                        # companies = [['name1','address1','tel1','web1'], ['name2','address2','tel2','web2']]

                        for info in data_to_write:
                            page.append(info)

                        wb.save(filename=workbook_name)

                    else:
                        # Make a new spreadsheet
                        
                        #['Name','RMSE','MAE','REL','d105', 'd110', 'd125']
                        workbook_name = os.path.join(data_path,subset+'.xlsx')
                        wb = Workbook()
                        page = wb.active
                        page.title = 'Info'
                        page.append(headers) # write the headers to the first line

                        # Data to write:
                        data_to_write = [[arrangement_path] + [float(metrics_meter.meters[header_name].stats.enum[-1]) for header_name in headers[1:]]]
                        # companies = [['name1','address1','tel1','web1'], ['name2','address2','tel2','web2']]


                        # Convert data to float if needed
                        for elem in data_to_write[0][1:]:
                            if type(elem) != float:
                                elem = float(elem.numpy())
                        for info in data_to_write:
                            page.append(info)
                        wb.save(filename = workbook_name)
                idx += 1



                # metrics_meter.update(post_pred.masked_fill(mask_bool,0), batch["gt_depth"].masked_fill(mask_bool,0))
                # metrics_meter.update(torch.matmul(post_pred,1-batch["mask"]), torch.matmul(batch["gt_depth"],1-batch["mask"]))
        print(np.average(np.array(rmse_temp)))
        state = "Inference: subset -- {} | ".format(subset)
        logger.info(state + metrics_meter.suffix)

# def inference(
#         model, test_loaders, metrics, save_dir="", logger=None
# ):

#     model.eval()
#     metrics_meter = AggregatedMeter(metrics, maxlen=20)
#     for subset, loader in test_loaders.items():
#         idx = 0
#         logger.info(
#             "Inference: subset -- {}. Total number of batches: {}.".format(subset, len(loader))
#         )

#         metrics_meter.reset()
#         # loop over dataset
#         for batch in tqdm(loader):
#             batch = model.preprocess(batch)
#             pred = model(batch)

#             with torch.no_grad():
#                 post_pred = model.postprocess(pred)
#                 if save_dir:
#                     B = batch["color"].shape[0]
#                     for it in range(B):
#                         fig = visualize.figure(
#                             batch["color"][it], batch["raw_depth"][it],
#                             batch["mask"][it], batch["gt_depth"][it],
#                             post_pred[it], close=True
#                         )
#                         fig.savefig(
#                             os.path.join(save_dir, "result_{}.png".format(idx)), dpi=fig.dpi
#                         )

#                         idx += 1

#                 metrics_meter.update(post_pred, batch["gt_depth"])

#         state = "Inference: subset -- {} | ".format(subset)
#         logger.info(state + metrics_meter.suffix)